# This script finds the best character-vehicle combination for a Mario Kart World race in 150cc with normal items.
# It achieves its objective by maximising its expected speed during a race while fulfilling a minimum acceleration requirement.
# The evaluation of the expected speed takes the effects of coins during races in terms of the probabilities of the numbers of coins the player is expected to possess and the combination's responsiveness to the coins.

import numpy as np
import openpyxl as pyxl
import Functions as f

## Initiate and preallocate

file=pyxl.load_workbook("Mario Kart World stats.xlsx",data_only=True)
# loads the excel file
characters=file["Characters"]
vehicles=file["Vehicles"]
weightandcoins=file["Weight & Coins"]
acceleration=file["Acceleration"]
# separates the spreadsheets in the excel files

nc=20
# 20 unique character classes
nv=24
# 24 unique vehicle stat configurations
nB=nc*nv
# total number of possible unique combinations
iB=0
# preallocates the considered combination counter

comboname=[[0]*1 for i2 in range(0,nB,1)]
speedlvl_solid=np.zeros(nB)
speedlvl_grainy=np.zeros(nB)
speedlvl_water=np.zeros(nB)
speedincrease_solid=np.zeros((nB,21))
speedincrease_grainy=np.zeros((nB,21))
speedincrease_water=np.zeros((nB,21))
speedincrease_railairorwall=np.zeros((nB,21))
speedincrease_coinaveraged=np.zeros((nB,21))
nonitemboost_rates=np.zeros(7)
nonitemboost_durations=np.zeros((nB,7))
nonitemboost_proportion=np.zeros(nB)
speedincrease_overall=np.zeros((nB,21))
speedincrease_expected=np.zeros(nB)
accelerationlevel=np.zeros(nB)
weight=np.zeros(nB)
handling_solid=np.zeros(nB)
handling_grainy=np.zeros(nB)
handling_water=np.zeros(nB)
# preallocates arrays to store the considered combinations

## Analyse the stats

# Chooses whether to use Queueing Theory probabilities (True) or from raw data (False)
if True:
    P=f.CoinCountProbabilities(True)
    # loads the state probabilities of the number of coins the player posseses at any point during a race
    # The boolean input indicates whether the focus is on 3-lap races (True) or 6-lap knockout tours (False)
else:
    file2=pyxl.load_workbook("Mario Kart World real time coin possession data.xlsx",data_only=True)
    spreadsheet=file2["shortcat"]
    # loads the raw data spreadsheet
    P=np.zeros(21)
    # preallocates the probability matrix
    for i in range(0,21,1):
        P[i]=spreadsheet.cell(row=2+i,column=7).value
        # imports the probabilities

file3=pyxl.load_workbook("Mario Kart World race data.xlsx",data_only=True)
nonitemboosts=file3["Sheet1"]
for ib in range(0,7,1):
    nonitemboost_rates[ib]=nonitemboosts.cell(row=4+ib,column=16).value
raceduration=60/(nonitemboosts.cell(row=11,column=16).value)
nonitemboost_speedincrease=acceleration.cell(row=4,column=26).value
# loads the non-item boost rates and durations

for ic in range(0,nc,1):
    for iv in range(0,nv,1):
        # for every character-vehicle combination
        comboname[iB]=f"{characters.cell(row=3+ic,column=1).value} with the {vehicles.cell(row=3+iv,column=1).value}"
        speedlvl_solid[iB]=characters.cell(row=3+ic,column=2).value+vehicles.cell(row=3+iv,column=2).value
        speedlvl_grainy[iB]=characters.cell(row=3+ic,column=3).value+vehicles.cell(row=3+iv,column=3).value
        speedlvl_water[iB]=characters.cell(row=3+ic,column=4).value+vehicles.cell(row=3+iv,column=4).value
        accelerationlevel[iB]=characters.cell(row=3+ic,column=5).value+vehicles.cell(row=3+iv,column=5).value
        weight[iB]=characters.cell(row=3+ic,column=6).value+vehicles.cell(row=3+iv,column=6).value
        handling_solid[iB]=characters.cell(row=3+ic,column=7).value+vehicles.cell(row=3+iv,column=7).value
        handling_grainy[iB]=characters.cell(row=3+ic,column=8).value+vehicles.cell(row=3+iv,column=8).value
        handling_water[iB]=characters.cell(row=3+ic,column=9).value+vehicles.cell(row=3+iv,column=9).value
        # records all relevant stat levels by the summation of character and vehicle stat levels
        for i in range(0,21,1):
            speedincrease_solid[iB,i]=(1+weightandcoins.cell(row=2+int(speedlvl_solid[iB]),column=2).value)*(1+weightandcoins.cell(row=2+int(weight[iB]),column=5+i).value)-1
            speedincrease_grainy[iB,i]=(1+weightandcoins.cell(row=2+int(speedlvl_grainy[iB]),column=2).value)*(1+weightandcoins.cell(row=2+int(weight[iB]),column=5+i).value)-1
            speedincrease_water[iB,i]=(1+weightandcoins.cell(row=2+int(speedlvl_water[iB]),column=2).value)*(1+weightandcoins.cell(row=2+int(weight[iB]),column=5+i).value)-1
            speedincrease_railairorwall[iB,i]=(1+weightandcoins.cell(row=15,column=2).value)*(1+weightandcoins.cell(row=2+int(weight[iB]),column=5+i).value)-1
            # recalls the speed increases for each terrain at every coin count according to the speed level and weight
        speedincrease_coinaveraged[iB,:]=(0.5188*speedincrease_solid[iB,:]+0.2131*speedincrease_grainy[iB,:]+0.0747*speedincrease_water[iB,:]+0.1934*speedincrease_railairorwall[iB,:])
        # evaluates the coin-averaged speed increase
        for ib in range(0,7,1):
            nonitemboost_durations[iB,ib]=acceleration.cell(row=1+int(accelerationlevel[iB]),column=4+3*ib).value
            # finds the non-item boost durations
        nonitemboost_proportion[iB]=np.dot(np.array(nonitemboost_rates).flatten()/60,np.array(nonitemboost_durations[iB,:]).flatten()).item()
        speedincrease_overall[iB,:]=((1+speedincrease_coinaveraged[iB,:])*(1+nonitemboost_proportion[iB]*nonitemboost_speedincrease))*(raceduration)/(raceduration+acceleration.cell(row=1+int(accelerationlevel[iB]),column=23).value-acceleration.cell(row=4,column=23).value)-1
        # finds the speed increases at every coin including the effects of non-item boosts
        speedincrease_expected[iB]=np.dot(speedincrease_overall[iB,:],P).item()
        # evaluates the overall speed increase by the proportions of solid, grainy, and water terrains
        # treats the number of coins possessed and the terrain on which the player is driving as independent events
        iB+=1
        # counts and saves the considered combination

## Find the best combination and viable alternatives

jB1=speedincrease_expected.argmax()

jB=np.array(np.where(speedincrease_expected>=np.max(speedincrease_expected)-0.1*0.01)).flatten()
# finds the indices of all viable alternatives (within 2% of the maximum relative expected speed increase)
jB=np.delete(jB,np.where(jB==jB1))
jB=np.concatenate((np.array([jB1]),jB[:]))
# moves the best combination to the top of the list

## find how a particular chosen combination compares (optinal; toggle by boolean)

user_selection=True

if user_selection:
    IC=10 # Swoop
    IV=3 # Reel racer
    IB=IC*nv+IV
    # the index of the chosen combination
    jB=np.concatenate((np.array([IB]),jB[:]))
    # adds the user-selected combination to the top of the list
    
## display the results and upload them onto the outfile

answers=pyxl.load_workbook("Answers.xlsx")
answersheet=answers["Sheet1"]
# opens the outfile

for ia in range(1,11,1):
    for i1 in range(3,483,1):
        cell1=answersheet.cell(row=i1,column=ia)
        cell1.value=None
for ia in range(1,481,1):
    for i1 in range(2,24,1):
        cell2=answersheet.cell(row=i1,column=13+ia)
        cell2.value=None
# clears all existing answers on the outfile

for c in range(0,21,1):
    answersheet.cell(row=3+c,column=13,value=np.array(P).flatten()[c])
    # uploads the probabilities onto the answer spreadsheet

if user_selection:
    s1=np.sort(np.array(speedincrease_expected).flatten()).copy()
    s1=s1[-1::-1]
    ranks=np.array(np.where(s1==speedincrease_expected[IB])).flatten()
    if ranks.size!=0:
        rank=ranks[0]+1
    else:
        rank=np.nan
    for j in range(0,len(jB),1):
        # for each considered character-vehicle combination
        answersheet.cell(row=2,column=14+j,value=j)
        answersheet.cell(row=3+j,column=2,value=comboname[jB[j]])
        answersheet.cell(row=3+j,column=3,value=int(speedlvl_solid[jB[j]]))
        answersheet.cell(row=3+j,column=4,value=int(speedlvl_grainy[jB[j]]))
        answersheet.cell(row=3+j,column=5,value=int(speedlvl_water[jB[j]]))
        answersheet.cell(row=3+j,column=6,value=int(accelerationlevel[jB[j]]))
        answersheet.cell(row=3+j,column=7,value=int(weight[jB[j]]))
        answersheet.cell(row=3+j,column=8,value=int(handling_solid[jB[j]]))
        answersheet.cell(row=3+j,column=9,value=int(handling_grainy[jB[j]]))
        answersheet.cell(row=3+j,column=10,value=int(handling_water[jB[j]]))
        # uploads the attribute profile onto the outfile
        for c in range(0,21,1):
            answersheet.cell(row=3+c,column=14+j,value=speedincrease_overall[jB[j],c])
            # uploads the overall speed increase at every coin onto the outfile

        if j==0:
            answersheet.cell(row=3+j,column=1,value="User-selected")
            answersheet.cell(row=2,column=14+j,value="User-selected")
            print(f"User-selected combination: {comboname[jB[j]]}")
        elif j==1:
            answersheet.cell(row=3+j,column=1,value=j)
            answersheet.cell(row=2,column=14+j,value=j)
            print("")
            print(f"Best combination: {comboname[jB[j]]}")
        else:
            answersheet.cell(row=3+j,column=1,value=j)
            answersheet.cell(row=2,column=14+j,value=j)
            print("")
            print(f"Viable alternative {j-1}: {comboname[jB[j]]}")

        print(f"Expected speed: +{np.round(100*speedincrease_expected[jB[j]],3):.3f}% from baseline")
        print(f"Speed level: {int(speedlvl_solid[jB[j]])}-{int(speedlvl_grainy[jB[j]])}-{int(speedlvl_water[jB[j]])}")
        print(f"Acceleration level: {int(accelerationlevel[jB[j]])}")
        print(f"Weight: {int(weight[jB[j]])}")
        print(f"Handling: {int(handling_solid[jB[j]])}-{int(handling_grainy[jB[j]])}-{int(handling_water[jB[j]])}")

        if j==0:
            if rank==1:
                print(f"This combination is the best")
            elif rank==2:
                print(f"This combination is the 2nd best")
            elif rank==3:
                print(f"This combination is the 3rd best")
            elif rank%10>=4 or rank%10==0 or (rank%100>=10 and rank%100<=20):
                print(f"This combination is the {rank}th best")
            elif rank%10==3:
                print(f"This combination is the {rank}rd best")
            elif rank%10==2:
                print(f"This combination is the {rank}nd best")
            elif rank%10==1:
                print(f"This combination is the {rank}st best")
            # prints the details of the selected combination
    
else:
    for j in range(0,len(jB),1):
        # for each considered character-vehicle combination
        answersheet.cell(row=4+j,column=1,value=j+1)
        answersheet.cell(row=2,column=15+j,value=j+1)
        answersheet.cell(row=4+j,column=2,value=comboname[jB[j]])
        answersheet.cell(row=4+j,column=3,value=int(speedlvl_solid[jB[j]]))
        answersheet.cell(row=4+j,column=4,value=int(speedlvl_grainy[jB[j]]))
        answersheet.cell(row=4+j,column=5,value=int(speedlvl_water[jB[j]]))
        answersheet.cell(row=4+j,column=6,value=int(accelerationlevel[jB[j]]))
        answersheet.cell(row=4+j,column=7,value=int(weight[jB[j]]))
        answersheet.cell(row=4+j,column=8,value=int(handling_solid[jB[j]]))
        answersheet.cell(row=4+j,column=9,value=int(handling_grainy[jB[j]]))
        answersheet.cell(row=4+j,column=10,value=int(handling_water[jB[j]]))
        # uploads the attribute profile onto the outfile
        for c in range(0,21,1):
            answersheet.cell(row=3+c,column=15+j,value=speedincrease_coinaveraged[jB[j],c])
            # uploads the overall speed increase at every coin onto the outfile

        if j==0:
            print(f"Best combination: {comboname[jB[j]]}")
        else:
            print("")
            print(f"Viable alternative {j}: {comboname[jB[j]]}")

        print(f"Expected speed: +{np.round(100*speedincrease_expected[jB[j]],3):.3f}% from baseline")
        print(f"Speed level: {int(speedlvl_solid[jB[j]])}-{int(speedlvl_grainy[jB[j]])}-{int(speedlvl_water[jB[j]])}")
        print(f"Acceleration level: {int(accelerationlevel[jB[j]])}")
        print(f"Weight: {int(weight[jB[j]])}")
        print(f"Handling: {int(handling_solid[jB[j]])}-{int(handling_grainy[jB[j]])}-{int(handling_water[jB[j]])}") 

answers.save("Answers.xlsx")
# saves the outfile
