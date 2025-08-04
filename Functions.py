import numpy as np
import openpyxl as pyxl

def CoinCountProbabilities(race_or_KO):
    # This function uses queueing theory to find the probabilities of all the coin counts from 0 to 20 during a race in Mario Kart World.
    # It requires a boolean input to indicate whether 3-lap races (True) or 6-lap knockout tours (False) are considered.

    file=pyxl.load_workbook("Mario Kart World race coin and crash data.xlsx",data_only=True)
    spreadsheet=file["Sheet1"]
    Lambda=spreadsheet.cell(row=2,column=9).value
    # The collection rate of coins (1/min)
    mu=spreadsheet.cell(row=3,column=9).value
    # The frequency of the player hit by an item and loses up to 3 coins each time (1/min)
    r=spreadsheet.cell(row=4,column=9).value*(3+0.5*race_or_KO)/3.5
    # The frequency of the player finishing the race and returning to the starting coin count between 0 and 5 (1/min)
    # The expected duration of a 6-lap knockout tour is around 3.5 laps (i.e. 1/6*6? = 21/6 = 3.5)

    A=np.zeros((21,21))

    A[0,:]=np.array([-Lambda]+[mu+5/6*r]*3+[5/6*r]*2+[r]*15)

    for i in range(1,6,1):
        A[i,:]=np.array([0]*(i-1)+[Lambda,-(Lambda+mu+5/6*r),0,0,mu]+[0]*(17-i))

    for i in range(6,18,1):
        A[i,:]=np.array([0]*(i-1)+[Lambda,-(Lambda+mu+r),0,0,mu]+[0]*(17-i))

    A[18,:]=np.array([0]*17+[Lambda,-(Lambda+mu+r),0,0])
    A[19,:]=np.array([0]*18+[Lambda,-(mu+r),-(mu+r)])
    A[20,:]=1
    # Sets up the queueing theory equations in vector-matrix form using the principle that the probability of arriving a state (coin count) is equal to that of leaving a state
    # The vector-matrix equation will be A*P=B, where A is a 21-by-21 matrix containing the coefficients, P is the column matrix of the state probabilities, and B is the column matrix of the 

    B=np.zeros((21,1))
    B[-1]=1

    P=np.linalg.lstsq(A,B,rcond=None)[0]
    # Finds the state probabilities of each coin count from 0 to 20

    return P
